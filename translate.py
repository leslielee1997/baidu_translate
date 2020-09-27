# coding=utf-8
import http.client
import hashlib
import urllib
import random
import json
import time
import openpyxl
from openpyxl import Workbook
def transbaidu(res_q):
    appid = '你的id'  # 百度翻译开发者appid
    secretKey = '你的密钥'  # 开发密钥
    httpClient = None
    myurl = '/api/trans/vip/translate'#url地址
    fromLang = 'auto'   #原文语种
    toLang = 'zh'   #译文语种
    salt = random.randint(32768, 65536)
    q= res_q  #翻译内容
    sign = appid + q + str(salt) + secretKey  #密文
    sign = hashlib.md5(sign.encode()).hexdigest() #加密MD5
    myurl = myurl + '?appid=' + appid + '&q=' + urllib.parse.quote(q) + '&from=' + fromLang + '&to=' + toLang + '&salt=' + str(
    salt) + '&sign=' + sign  #最终url

    try:
        httpClient = http.client.HTTPConnection('api.fanyi.baidu.com')
        httpClient.request('GET', myurl)
        # response是HTTPResponse对象
        response = httpClient.getresponse()
        result_all = response.read().decode("utf-8")
        result = json.loads(result_all)
        #print (result['trans_result'][0]['src'])原文
        res=result['trans_result'][0]['dst']
        return res
    except Exception as e:
        print (e)
    finally:
        if httpClient:
            httpClient.close()
            
            
def excelTrans(
        srcFilename=r'D:\test\source.xlsx',
        desFilename=r'D:\test\result.xlsx',
        srcSheet='Sheet1',        
        num = 2,
        #srcColumn=2,
        srcRowBegin=1,
        srcRowEnd=44,
        desColumn=1,
        desSheet='result2'):
    wb = openpyxl.load_workbook(srcFilename)
    ws = wb[srcSheet]
    wb2 = Workbook()
    #ws2 = wb2.create_sheet(title=desSheet)    
    #ws2 = wb2.create_sheet(title=desSheet,index = 1)
    for j in range(num):        
        ws2 = wb2.create_sheet(title=desSheet,index = j)        
        for i in range(srcRowBegin, srcRowEnd, 1):
            sstr = ws.cell(row=i, column=j+1).value
            if not (sstr is None):
                ws2.cell(row=i-srcRowBegin+1, column=desColumn).value = transbaidu(sstr)
                time.sleep(1)  #反爬，设置定时；数据太大时就要用多线程了。
        wb2.save(desFilename)

if __name__ == '__main__':
    excelTrans()
